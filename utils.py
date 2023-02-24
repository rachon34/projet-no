import pandas as pd
import openpyxl
from data_visualizer import create_dashboard

from config import Config
from config import Config
DATA_FILE = Config.DATA_FILE
MIN_SLEEP = Config.MIN_SLEEP
MAX_SLEEP = Config.MAX_SLEEP



...
import subprocess

def install_missing_packages(packages):
    """Installe automatiquement les paquets manquants.

    Args:
        packages (list): Liste des paquets à installer.

    Returns:
        int: Code de retour de la commande pip.
    """
    # Liste des paquets manquants
    missing_packages = []
    for package in packages:
        try:
            __import__(package)
        except ImportError:
            missing_packages.append(package)

    # Installation des paquets manquants
    if missing_packages:
        print("Les paquets suivants sont manquants :")
        print("\n".join(missing_packages))
        print("Installation en cours...")

        command = ['pip', 'install'] + missing_packages
        return subprocess.call(command)
    else:
        print("Toutes les dépendances sont déjà installées.")
        return 0

def add_entry_to_excel(entry):
    wb = openpyxl.load_workbook(DATA_FILE)
    sheet = wb.active
    sheet.append([entry['name'], entry['description'], entry['price']])
    wb.save(DATA_FILE)

def delete_entry_from_excel(entry):
    wb = openpyxl.load_workbook(DATA_FILE)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        if row[0] == entry['name'] and row[1] == entry['description'] and row[2] == entry['price']:
            sheet.delete_rows(row[0].row)
            break
    wb.save(DATA_FILE)

def get_entry_by_id(id):
    wb = openpyxl.load_workbook(DATA_FILE)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == id:
            return {'name': row[1], 'description': row[2], 'price': row[3]}
    return None

import pandas as pd


def add_entry_to_excel(data, file_path):
    df = pd.read_excel(file_path)
    df = df.append(data, ignore_index=True)
    df.to_excel(file_path, index=False)


def delete_entry_from_excel(entry_id, file_path):
    df = pd.read_excel(file_path)
    df = df[df['id'] != entry_id]
    df.to_excel(file_path, index=False)


def get_entry_by_id(entry_id, file_path):
    df = pd.read_excel(file_path)
    entry = df.loc[df['id'] == entry_id].to_dict(orient='records')
    return entry[0] if entry else None
